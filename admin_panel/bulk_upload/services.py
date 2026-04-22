"""
Bulk upload: parse CSV/XLSX, validate rows, create users and profile rows.
"""
from __future__ import annotations

import csv
import io
import re
import secrets
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.validators import URLValidator, validate_email
from django.db import transaction
from django.db.models import Q

from accounts.models import User
from master.models import (
    Caste,
    City,
    Country,
    District,
    Education,
    EducationSubject,
    Height,
    IncomeRange,
    MaritalStatus,
    MotherTongue,
    Occupation,
    Religion,
    State,
)
from profiles.models import (
    UserEducation,
    UserFamily,
    UserLocation,
    UserPersonal,
    UserPhotos,
    UserProfile,
    UserReligion,
)
from profiles.utils import get_profile_completion_data

CACHE_PREFIX = "bulk_upload:"
CACHE_TTL = 60 * 60  # 1 hour
ASYNC_ROW_THRESHOLD = 50

TEMPLATE_COLUMNS = [
    "Name",
    "Age",
    "Gender",
    "Date of Birth",
    "Mobile Number",
    "Email",
    "Country",
    "State",
    "District",
    "City",
    "Religion",
    "Caste",
    "Mother Tongue",
    "Gothram",
    "Marital Status",
    "Height (cm)",
    "Weight (kg)",
    "Complexion",
    "No. of Children",
    "Children Living With",
    "Education",
    "Education Subject",
    "Occupation",
    "Company",
    "Annual Income",
    "Working Location",
    "About Me",
    "Father Name",
    "Father Occupation",
    "Mother Name",
    "Mother Occupation",
    "No. of Brothers",
    "No. of Sisters",
    "Family Type",
    "Family Status",
    "Rasi",
    "Nakshatra",
    "Lagnam",
    "Dosham",
    "Birth Time",
    "Birth Place",
    "Photo URL",
]

# Normalized header label -> internal key
_LABEL_TO_KEY: dict[str, str] = {
    "name": "name",
    "age": "age",
    "gender": "gender",
    "date of birth": "dob",
    "mobile number": "mobile",
    "email": "email",
    "country": "country",
    "state": "state",
    "district": "district",
    "city": "city",
    "religion": "religion",
    "caste": "caste",
    "mother tongue": "mother_tongue",
    "gothram": "gothram",
    "marital status": "marital_status",
    "height (cm)": "height_cm",
    "weight (kg)": "weight_kg",
    "complexion": "complexion",
    "no. of children": "num_children",
    "children living with": "children_living_with",
    "education": "education",
    "education subject": "education_subject",
    "occupation": "occupation",
    "company": "company",
    "annual income": "annual_income",
    "working location": "working_location",
    "about me": "about_me",
    "father name": "father_name",
    "father occupation": "father_occupation",
    "mother name": "mother_name",
    "mother occupation": "mother_occupation",
    "no. of brothers": "num_brothers",
    "no. of sisters": "num_sisters",
    "family type": "family_type",
    "family status": "family_status",
    "rasi": "rasi",
    "nakshatra": "nakshatra",
    "lagnam": "lagnam",
    "dosham": "dosham",
    "birth time": "birth_time",
    "birth place": "birth_place",
    "photo url": "photo_url",
}


def _norm_label(cell: Any) -> str:
    if cell is None:
        return ""
    return " ".join(str(cell).strip().split()).lower()


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
    return str(val).strip()


def parse_spreadsheet_file(uploaded_file) -> tuple[list[str], list[dict[str, str]]]:
    """Return (header_labels, rows) where each row is label->string value."""
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
    if name.endswith(".csv") or not name.endswith(".xlsx"):
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        return _parse_csv_text(text)
    return _parse_xlsx_bytes(raw)


def _parse_csv_text(text: str) -> tuple[list[str], list[dict[str, str]]]:
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_norm_label(h) for h in header]
    data: list[dict[str, str]] = []
    for row in rows_iter:
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            key = _LABEL_TO_KEY.get(h)
            if not key:
                continue
            d[key] = _cell_str(row[i]) if i < len(row) else ""
        data.append(d)
    return headers, data


def _parse_xlsx_bytes(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_norm_label(h) for h in header_row]
    data: list[dict[str, str]] = []
    for row in rows_iter:
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            key = _LABEL_TO_KEY.get(h)
            if not key:
                continue
            val = row[i] if i < len(row) else None
            if val is not None and not isinstance(val, str):
                if hasattr(val, "strftime"):
                    d[key] = val.strftime("%d-%m-%Y")
                else:
                    d[key] = _cell_str(val)
            else:
                d[key] = _cell_str(val)
        data.append(d)
    wb.close()
    return headers, data


def _row_is_empty(row: dict[str, str]) -> bool:
    return not any((v or "").strip() for v in row.values())


def normalize_mobile(raw: str) -> str | None:
    if raw is None:
        return None
    digits = "".join(c for c in str(raw) if c.isdigit())
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[-10:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[-10:]
    if len(digits) == 10:
        return f"+91{digits}"
    return None


def mobile_exists_in_db(mobile_e164: str) -> bool:
    digits = "".join(c for c in str(mobile_e164) if c.isdigit())
    if len(digits) == 12 and digits.startswith("91"):
        mobile_10 = digits[-10:]
    elif len(digits) == 10:
        mobile_10 = digits
    else:
        return False
    return User.objects.filter(
        Q(mobile=mobile_10)
        | Q(mobile="+91" + mobile_10)
        | Q(mobile="91" + mobile_10)
    ).exists()


def _parse_dob_dd_mm_yyyy(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d-%m-%Y").date()
    except ValueError:
        return None


def _resolve_master_name(model, name: str):
    if not (name or "").strip():
        return None
    n = name.strip()
    variants = [n, n + ".", n.rstrip(".")]
    for v in variants:
        obj = model.objects.filter(name__iexact=v, is_active=True).first()
        if obj:
            return obj
    return None


def _resolve_location(country_s, state_s, district_s, city_s):
    country = _resolve_master_name(Country, country_s) if country_s else None
    state = None
    if state_s and country:
        state = State.objects.filter(country=country, name__iexact=state_s.strip(), is_active=True).first()
    district = None
    if district_s and state:
        district = District.objects.filter(state=state, name__iexact=district_s.strip(), is_active=True).first()
    city = None
    if city_s and district:
        city = City.objects.filter(district=district, name__iexact=city_s.strip(), is_active=True).first()
    return country, state, district, city


def _get_or_create_height_cm(cm: int) -> Height:
    h, _ = Height.objects.get_or_create(
        value_cm=cm,
        defaults={"display_label": f"{cm} cm", "is_active": True},
    )
    return h


def _validate_email_optional(s: str) -> str | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        validate_email(s)
    except ValidationError:
        raise ValueError("invalid")
    return s


def _validate_url_optional(s: str) -> str | None:
    s = (s or "").strip()
    if not s:
        return None
    URLValidator()(s)
    return s


def validate_rows(
    data_rows: list[dict[str, str]],
) -> tuple[int, int, int, list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Returns total_rows, valid_rows, error_rows, errors list, valid import payloads.
    Row numbers are 1-based file row indices (first data row = 2 if header is row 1).
    """
    errors: list[dict[str, Any]] = []
    valid_payloads: list[dict[str, Any]] = []

    # First pass: row numbers for non-empty rows
    non_empty: list[tuple[int, dict[str, str]]] = []
    row_num = 2  # assume header at 1
    for row in data_rows:
        if _row_is_empty(row):
            row_num += 1
            continue
        non_empty.append((row_num, row))
        row_num += 1

    total_rows = len(non_empty)
    seen_mobile: dict[str, int] = {}

    for row_index, row in non_empty:
        row_errors: list[dict[str, Any]] = []

        name = (row.get("name") or "").strip()
        if not name:
            row_errors.append(
                {"row": row_index, "field": "name", "message": "Name is required"}
            )

        mobile_raw = row.get("mobile") or ""
        mobile = normalize_mobile(mobile_raw)
        if not (mobile_raw or "").strip():
            row_errors.append(
                {"row": row_index, "field": "mobile", "message": "Invalid mobile"}
            )
        elif mobile is None:
            row_errors.append(
                {"row": row_index, "field": "mobile", "message": "Invalid mobile"}
            )
        elif mobile in seen_mobile:
            row_errors.append(
                {
                    "row": row_index,
                    "field": "mobile",
                    "message": "Duplicate mobile in upload file",
                }
            )
        else:
            if mobile_exists_in_db(mobile):
                row_errors.append(
                    {
                        "row": row_index,
                        "field": "mobile",
                        "message": "Mobile already registered",
                    }
                )
            seen_mobile[mobile] = row_index

        gender_raw = (row.get("gender") or "").strip().upper()
        if not gender_raw:
            row_errors.append(
                {"row": row_index, "field": "gender", "message": "Gender must be M, F, or O"}
            )
        elif gender_raw not in ("M", "F", "O"):
            row_errors.append(
                {"row": row_index, "field": "gender", "message": "Gender must be M, F, or O"}
            )

        dob = _parse_dob_dd_mm_yyyy(row.get("dob") or "")
        if (row.get("dob") or "").strip() and dob is None:
            row_errors.append(
                {
                    "row": row_index,
                    "field": "dob",
                    "message": "Invalid date format, use DD-MM-YYYY",
                }
            )

        religion_name = (row.get("religion") or "").strip()
        religion_obj = None
        if religion_name:
            religion_obj = Religion.objects.filter(name__iexact=religion_name, is_active=True).first()
            if religion_obj is None:
                row_errors.append(
                    {
                        "row": row_index,
                        "field": "religion",
                        "message": f"Religion '{religion_name}' not found",
                    }
                )

        caste_name = (row.get("caste") or "").strip()
        caste_obj = None
        if caste_name:
            if not religion_name:
                row_errors.append(
                    {"row": row_index, "field": "religion", "message": "Religion is required"}
                )
            elif religion_obj:
                caste_obj = Caste.objects.filter(
                    religion=religion_obj, name__iexact=caste_name, is_active=True
                ).first()
                if caste_obj is None:
                    row_errors.append(
                        {
                            "row": row_index,
                            "field": "caste",
                            "message": f"Caste '{caste_name}' not found for religion '{religion_obj.name}'",
                        }
                    )

        height_raw = (row.get("height_cm") or "").strip()
        height_cm: int | None = None
        if height_raw:
            try:
                height_cm = int(float(height_raw))
            except (ValueError, TypeError):
                height_cm = None
            if height_cm is None or not (100 <= height_cm <= 250):
                row_errors.append(
                    {
                        "row": row_index,
                        "field": "height",
                        "message": "Height must be between 100 and 250 cm",
                    }
                )

        email_val = (row.get("email") or "").strip()
        if email_val:
            try:
                _validate_email_optional(email_val)
            except ValueError:
                row_errors.append(
                    {
                        "row": row_index,
                        "field": "email",
                        "message": "Invalid email address",
                    }
                )

        errors.extend(row_errors)

        if row_errors:
            continue

        # Build resolved payload (JSON-serializable)
        email_clean = _validate_email_optional(email_val) if email_val else None
        country, state, district, city = _resolve_location(
            (row.get("country") or "").strip(),
            (row.get("state") or "").strip(),
            (row.get("district") or "").strip(),
            (row.get("city") or "").strip(),
        )

        mt = _resolve_master_name(MotherTongue, (row.get("mother_tongue") or "").strip())
        ms = _resolve_master_name(MaritalStatus, (row.get("marital_status") or "").strip())
        edu = _resolve_master_name(Education, (row.get("education") or "").strip())
        edus = _resolve_master_name(EducationSubject, (row.get("education_subject") or "").strip())
        occ = _resolve_master_name(Occupation, (row.get("occupation") or "").strip())
        inc = _resolve_master_name(IncomeRange, (row.get("annual_income") or "").strip())

        weight_dec = None
        wr = (row.get("weight_kg") or "").strip()
        if wr:
            try:
                weight_dec = Decimal(str(wr))
            except (InvalidOperation, ValueError):
                weight_dec = None

        num_children = 0
        nc = (row.get("num_children") or "").strip()
        if nc:
            try:
                num_children = max(0, int(float(nc)))
            except (ValueError, TypeError):
                num_children = 0

        nb = (row.get("num_brothers") or "").strip()
        ns = (row.get("num_sisters") or "").strip()
        try:
            brothers = max(0, int(float(nb))) if nb else 0
        except (ValueError, TypeError):
            brothers = 0
        try:
            sisters = max(0, int(float(ns))) if ns else 0
        except (ValueError, TypeError):
            sisters = 0

        height_fk_id = None
        if height_cm is not None:
            height_fk_id = _get_or_create_height_cm(height_cm).pk

        horoscope_data = {}
        for fld in ("rasi", "nakshatra", "lagnam", "dosham", "birth_time", "birth_place"):
            v = (row.get(fld) or "").strip()
            if v:
                horoscope_data[fld] = v

        photo_url = None
        pur = (row.get("photo_url") or "").strip()
        if pur:
            try:
                photo_url = _validate_url_optional(pur)
            except ValidationError:
                photo_url = None

        complexion = (row.get("complexion") or "").strip()

        payload = {
            "row": row_index,
            "name": name,
            "mobile": mobile,
            "email": email_clean,
            "gender": gender_raw,
            "dob": dob.isoformat() if dob else None,
            "height_cm": height_cm,
            "religion_id": religion_obj.pk if religion_obj else None,
            "caste_id": caste_obj.pk if caste_obj else None,
            "mother_tongue_id": mt.pk if mt else None,
            "gothram": (row.get("gothram") or "").strip(),
            "marital_status_id": ms.pk if ms else None,
            "height_id": height_fk_id,
            "weight": str(weight_dec) if weight_dec is not None else None,
            "complexion": complexion,
            "num_children": num_children,
            "children_living_with": (row.get("children_living_with") or "").strip(),
            "education_id": edu.pk if edu else None,
            "education_subject_id": edus.pk if edus else None,
            "occupation_id": occ.pk if occ else None,
            "company": (row.get("company") or "").strip(),
            "annual_income_id": inc.pk if inc else None,
            "working_location": (row.get("working_location") or "").strip(),
            "about_me": (row.get("about_me") or "").strip(),
            "father_name": (row.get("father_name") or "").strip(),
            "father_occupation": (row.get("father_occupation") or "").strip(),
            "mother_name": (row.get("mother_name") or "").strip(),
            "mother_occupation": (row.get("mother_occupation") or "").strip(),
            "brothers": brothers,
            "sisters": sisters,
            "family_type": (row.get("family_type") or "").strip(),
            "family_status": (row.get("family_status") or "").strip(),
            "horoscope_data": horoscope_data,
            "photo_url": photo_url,
            "country_id": country.pk if country else None,
            "state_id": state.pk if state else None,
            "district_id": district.pk if district else None,
            "city_id": city.pk if city else None,
        }
        valid_payloads.append(payload)

    error_row_nums = {e["row"] for e in errors}
    error_rows = len(error_row_nums)
    valid_rows = total_rows - error_rows
    return total_rows, valid_rows, error_rows, errors, valid_payloads


def cache_validation_token(admin_user_id: int, rows: list[dict[str, Any]]) -> str:
    token = secrets.token_urlsafe(32)
    cache.set(
        CACHE_PREFIX + token,
        {"admin_user_id": admin_user_id, "rows": rows},
        CACHE_TTL,
    )
    return token


def get_cached_import(token: str) -> dict[str, Any] | None:
    return cache.get(CACHE_PREFIX + token)


def delete_cached_import(token: str) -> None:
    cache.delete(CACHE_PREFIX + token)


def import_profile_row(payload: dict[str, Any], branch_id: int | None) -> None:
    """Create User + profile records from a validated payload."""
    def _first(*keys, default=None):
        for k in keys:
            if k in payload and payload.get(k) not in (None, ""):
                return payload.get(k)
        return default

    dob = None
    if payload.get("dob"):
        dob = date.fromisoformat(payload["dob"])

    pwd = User.objects.make_random_password()
    email = payload.get("email") or None
    mobile = _first("mobile", "phone")
    user = User.objects.create_user(
        email=email,
        mobile=mobile,
        password=pwd,
        name=payload.get("name") or "",
        dob=dob,
        gender=payload.get("gender") or "",
        branch_id=branch_id,
    )
    user.is_active = True
    user.mobile_verified = True
    if email:
        user.email_verified = True
    # Ensure this is a member profile for admin panel lists.
    user.role = "user"
    user.save(
        update_fields=[
            "is_active",
            "mobile_verified",
            "email_verified",
            "role",
            "updated_at",
        ]
    )

    has_horo = bool(payload.get("horoscope_data"))
    profile, _ = UserProfile.objects.update_or_create(
        user=user,
        defaults={
            "about_me": payload.get("about_me") or "",
            "horoscope_data": payload.get("horoscope_data") or {},
            "has_horoscope": has_horo,
        },
    )

    UserLocation.objects.update_or_create(
        user=user,
        defaults={
            "country_id": payload.get("country_id"),
            "state_id": payload.get("state_id"),
            "district_id": payload.get("district_id"),
            "city_id": payload.get("city_id"),
            "address": payload.get("address") or "",
        },
    )

    ur_defaults = {
        "religion_id": payload.get("religion_id"),
        "caste_fk_id": payload.get("caste_id"),
        "mother_tongue_id": payload.get("mother_tongue_id"),
        "gothram": payload.get("gothram") or "",
    }
    if payload.get("caste_id"):
        from master.models import Caste as CasteModel

        c = CasteModel.objects.filter(pk=payload["caste_id"]).first()
        if c:
            ur_defaults["caste"] = c.name
    UserReligion.objects.update_or_create(user=user, defaults=ur_defaults)

    # children keys vary by validator version
    num_ch = int(_first("number_of_children", "num_children", default=0) or 0)
    weight = None
    weight_raw = _first("weight_kg", "weight")
    if weight_raw is not None:
        try:
            weight = Decimal(str(weight_raw))
        except (InvalidOperation, ValueError):
            weight = None

    h_cm = payload.get("height_cm")
    height_text = f"{h_cm} cm" if h_cm is not None else ""

    UserPersonal.objects.update_or_create(
        user=user,
        defaults={
            "marital_status_id": payload.get("marital_status_id"),
            "has_children": bool(payload.get("has_children")) if payload.get("has_children") is not None else num_ch > 0,
            "number_of_children": num_ch,
            "children_living_with": payload.get("children_living_with") or "",
            "height_id": payload.get("height_id"),
            "height_text": height_text,
            "weight": weight,
            "colour": payload.get("complexion") or "",
            "blood_group": payload.get("blood_group") or "",
        },
    )

    brothers = int(_first("brothers", "num_brothers", default=0) or 0)
    sisters = int(_first("sisters", "num_sisters", default=0) or 0)
    married_brothers = int(_first("married_brothers", "num_married_brothers", default=0) or 0)
    married_sisters = int(_first("married_sisters", "num_married_sisters", default=0) or 0)

    UserFamily.objects.update_or_create(
        user=user,
        defaults={
            "father_name": payload.get("father_name") or "",
            "father_occupation": payload.get("father_occupation") or "",
            "mother_name": payload.get("mother_name") or "",
            "mother_occupation": payload.get("mother_occupation") or "",
            "brothers": brothers,
            "sisters": sisters,
            "married_brothers": married_brothers,
            "married_sisters": married_sisters,
            "about_family": payload.get("about_family") or "",
            "family_type": payload.get("family_type") or "",
            "family_status": payload.get("family_status") or "",
        },
    )

    UserEducation.objects.update_or_create(
        user=user,
        defaults={
            "highest_education_id": _first("highest_education_id", "education_id"),
            "education_subject_id": payload.get("education_subject_id"),
            "occupation_id": payload.get("occupation_id"),
            "annual_income_id": payload.get("annual_income_id"),
            "employment_status": _first("employment", "employment_status", default="") or "",
            "company": payload.get("company") or "",
            "working_location": payload.get("working_location") or "",
        },
    )

    photos_defaults = {}
    if payload.get("photo_url"):
        photos_defaults["profile_photo_url"] = payload["photo_url"]
    UserPhotos.objects.update_or_create(user=user, defaults=photos_defaults)

    completion = get_profile_completion_data(user)
    user.is_registration_profile_completed = completion["profile_status"] == "completed"
    user.save(update_fields=["is_registration_profile_completed", "updated_at"])


def run_bulk_import(
    token: str, admin_user_id: int, branch_id: int | None
) -> dict[str, Any]:
    data = get_cached_import(token)
    if not data:
        return {"ok": False, "error": "Invalid or expired validation token"}
    if int(data.get("admin_user_id")) != int(admin_user_id):
        return {"ok": False, "error": "Validation token does not match this admin"}

    rows: list[dict[str, Any]] = data.get("rows") or []
    imported = 0
    failed: list[dict[str, Any]] = []

    for payload in rows:
        try:
            with transaction.atomic():
                import_profile_row(payload, branch_id)
            imported += 1
        except Exception as exc:
            failed.append({"row": payload.get("row"), "message": str(exc)})

    delete_cached_import(token)
    return {"ok": True, "imported": imported, "failed": failed}
