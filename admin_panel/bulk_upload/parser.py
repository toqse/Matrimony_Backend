import csv
import io
from typing import Any

from openpyxl import load_workbook

TEMPLATE_COLUMNS = [
    "Name",
    "Phone Number",
    "Email",
    "Date of Birth",
    "Gender",
    "Partner Preference",
    "Country",
    "State",
    "District",
    "City",
    "Address",
    "Religion",
    "Caste",
    "Mother Tongue",
    "Marital Status",
    "Has Children",
    "Number of Children",
    "Height (cm)",
    "Weight (kg)",
    "Complexion",
    "Highest Education",
    "Education Subject",
    "Employment",
    "Occupation",
    "Annual Income",
    "About Me",
    "Family Type",
    "Father's Name",
    "Father's Occupation",
    "Mother's Name",
    "Mother's Occupation",
    "Family Status",
    "Number of Brothers (integer >= 0)",
    "Number of Married Brothers (integer >= 0)",
    "Number of Sisters (integer >= 0)",
    "Number of Married Sisters (integer >= 0)",
    "About My Family",
]

_HEADER_TO_KEY = {
    "name": "name",
    "phone number": "phone",
    "email": "email",
    "date of birth": "dob",
    "gender": "gender",
    "partner preference": "partner_preference",
    "country": "country",
    "state": "state",
    "district": "district",
    "city": "city",
    "address": "address",
    "religion": "religion",
    "caste": "caste",
    "mother tongue": "mother_tongue",
    "marital status": "marital_status",
    "has children": "has_children",
    "number of children": "number_of_children",
    "height (cm)": "height_cm",
    "weight (kg)": "weight_kg",
    "complexion": "complexion",
    "highest education": "highest_education",
    "education subject": "education_subject",
    "employment": "employment",
    "occupation": "occupation",
    "annual income": "annual_income",
    "about me": "about_me",
    "family type": "family_type",
    "father's name": "father_name",
    "father name": "father_name",
    "father's status": "father_status",
    "father status": "father_status",
    "father's occupation": "father_occupation",
    "father occupation": "father_occupation",
    "mother's name": "mother_name",
    "mother name": "mother_name",
    "mother's status": "mother_status",
    "mother status": "mother_status",
    "mother's occupation": "mother_occupation",
    "mother occupation": "mother_occupation",
    "family status": "family_status",
    "number of brothers (integer >= 0)": "num_brothers",
    "number of married brothers (integer >= 0)": "num_married_brothers",
    "number of sisters (integer >= 0)": "num_sisters",
    "number of married sisters (integer >= 0)": "num_married_sisters",
    "about my family": "about_family",
}

# Excel often truncates headers; map normalized short headers when unambiguous.
_HEADER_ALIASES = {
    "phone nu": "phone",
    "phone no": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    "mother to": "mother_tongue",
    "marital sta": "marital_status",
    "partner pr": "partner_preference",
    "highest ed": "highest_education",
    "annual inc": "annual_income",
    "about my fam": "about_family",
    # Common shortened headings (without "(integer >= 0)" suffix)
    "number of brothers": "num_brothers",
    "number of married brothers": "num_married_brothers",
    "number of sisters": "num_sisters",
    "number of married sisters": "num_married_sisters",
    "about my family": "about_family",
    "about my family..": "about_family",
}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Excel stores phone counts etc. as float; avoid "9876543210.0" and scientific str quirks.
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
    return str(v).strip()


# Columns that uniquely identify the legacy matrimony export
# (typo `mother toungue`, plus the horoscope chart fields). When any of these
# is present we route the file through the legacy importer instead of
# rejecting it as "Unknown columns".
LEGACY_FORMAT_FINGERPRINTS = {
    "mother toungue",
    "horochart",
    "amsachart",
    "bhavchart",
    "sishta_dur",
    "padam",
}


def is_legacy_format(headers: list[str]) -> bool:
    """True when the header row matches the old matrimony CSV export."""
    normalized = {(h or "").strip().lower() for h in headers}
    return bool(LEGACY_FORMAT_FINGERPRINTS & normalized)


def parse_upload_file(uploaded_file) -> tuple[list[str], list[dict[str, str]]]:
    """Parse and strictly validate a bulk-upload template file.

    For modern templates this enforces exact header layout. Legacy matrimony
    exports are routed via parse_legacy_upload_file (skipping this validator).
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()
    data = uploaded_file.read()
    if name.endswith(".xlsx"):
        headers, rows = _parse_xlsx(data)
    else:
        headers, rows = _parse_csv(data)

    _validate_headers_strict(headers)
    return headers, rows


def read_upload_file_text(uploaded_file) -> tuple[str, list[str]]:
    """Decode an uploaded CSV/XLSX into CSV text + normalized header list.

    Used by the legacy-format path to keep the raw bytes for re-processing
    during import. XLSX is converted to CSV text on the fly so legacy
    detection and the legacy importer share a single text-based path.
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()
    data = uploaded_file.read()
    if name.endswith(".xlsx"):
        return _xlsx_to_csv_text(data)
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    try:
        first = next(reader)
    except StopIteration:
        return text, []
    return text, [_norm(h) for h in first]


def _xlsx_to_csv_text(raw: bytes) -> tuple[str, list[str]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    out = io.StringIO()
    writer = csv.writer(out)
    headers: list[str] = []
    for index, row in enumerate(ws.iter_rows(values_only=True)):
        cells = []
        for v in row:
            if hasattr(v, "strftime"):
                cells.append(v.strftime("%d/%m/%Y"))
            elif v is None:
                cells.append("")
            else:
                cells.append(str(v))
        writer.writerow(cells)
        if index == 0:
            headers = [_norm(h) for h in cells]
    wb.close()
    return out.getvalue(), headers


def _validate_headers_strict(headers: list[str]) -> None:
    """
    Enforce exact column mapping.
    - Normalization is already applied via _norm().
    - If a header isn't recognized -> error
    - If a required template header is missing -> error
    """
    provided = [h for h in headers if (h or "").strip()]
    provided_keys = [(_HEADER_TO_KEY.get(h) or _HEADER_ALIASES.get(h)) for h in provided]
    unknown = [h for h in provided if not (_HEADER_TO_KEY.get(h) or _HEADER_ALIASES.get(h))]

    required_headers = [_norm(h) for h in TEMPLATE_COLUMNS]
    required_keys = [(_HEADER_TO_KEY.get(h) or _HEADER_ALIASES.get(h)) for h in required_headers]
    required_keys = [k for k in required_keys if k]  # safety
    missing_keys = [k for k in required_keys if k not in set(provided_keys)]

    if unknown or missing_keys:
        parts: list[str] = []
        if missing_keys:
            parts.append("Missing columns: " + ", ".join(missing_keys))
        if unknown:
            parts.append("Unknown columns: " + ", ".join(unknown))
        raise ValueError("Invalid template headers. " + " | ".join(parts))


def _parse_csv(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [_norm(h) for h in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        item: dict[str, str] = {}
        for i, h in enumerate(headers):
            key = _HEADER_TO_KEY.get(h) or _HEADER_ALIASES.get(h)
            if not key:
                continue
            item[key] = _cell(row[i]) if i < len(row) else ""
        out.append(item)
    return headers, out


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        first = next(rows)
    except StopIteration:
        wb.close()
        return [], []
    headers = [_norm(h) for h in first]
    out: list[dict[str, str]] = []
    for row in rows:
        item: dict[str, str] = {}
        for i, h in enumerate(headers):
            key = _HEADER_TO_KEY.get(h) or _HEADER_ALIASES.get(h)
            if not key:
                continue
            val = row[i] if i < len(row) else None
            if hasattr(val, "strftime"):
                item[key] = val.strftime("%d-%m-%Y")
            else:
                item[key] = _cell(val)
        out.append(item)
    wb.close()
    return headers, out
